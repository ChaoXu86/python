# coding: utf-8
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

from optparse import OptionParser

COL_HEADER_ID = 3 # colum header row id
debugLevel = 'INFO'
 
class logging(object):
    def __init__(self,level):
        pass
        
    def __call__(self, func): 
        def wrapper(*args, **kwargs):                    
            if debugLevel == 'INFO':
                print("[{level}] [{time}]: enter function {func}()".format(level=debugLevel,time=datetime.datetime.now(),func=func.__name__))
            return func(*args, **kwargs)
        return wrapper  

@logging(level=debugLevel)
def load_xls(fileName):
    # get first worksheet
    wb = load_workbook(fileName)
    ws = wb[wb.sheetnames[0]]
    
    return wb,ws

@logging(level=debugLevel)
def filter_by_date(worksheet, timeStart, timeEnd = datetime.datetime.now()):    
    for date_col_id in range(worksheet.max_column):                
        if worksheet[COL_HEADER_ID][date_col_id].value == 'Created':
            break
    
    total_delete = 0
    row2delete_count = 0
    row2delete_from = -1
    for row_id in range(worksheet.max_row,COL_HEADER_ID,-1):   
        # delete rows from bottom to top
        ideaDate = worksheet[row_id][date_col_id].value        
        if ideaDate < timeStart or ideaDate > timeEnd:
            #print(row_id," row2delete: ", row2delete_from, " row2delete_count: ", row2delete_count)
            if row2delete_from == -1:                
                row2delete_from = row_id
            row2delete_count +=1
        else:
            # found first row that match the critera            
            if row2delete_count > 0:
                #print("delete from: ", row2delete_from - row2delete_count, ",count: ", row2delete_count)
                assert((row2delete_from - row2delete_count) > COL_HEADER_ID)
                worksheet.delete_rows(row2delete_from - row2delete_count, row2delete_count)
                #print("delete complete")
                total_delete += row2delete_count
                
                row2delete_count = 0  
                row2delete_from = -1    
    
    if row2delete_count > 0:
        #print("delete from: ", row2delete_from - row2delete_count, ",count: ", row2delete_count)
        worksheet.delete_rows(row2delete_from - row2delete_count, row2delete_count)
        total_delete += row2delete_count

    # remove first two useless rows
    worksheet.delete_rows(1, COL_HEADER_ID - 1)
        
    return total_delete

@logging(level=debugLevel)
def save(fileName, wb):    
    wb.save(fileName)

@logging(level=debugLevel)
def parse_idea_box(srcFileName,dstFileName,dateFrom,dateTo):
    wb,ws = load_xls(srcFileName)
    filter_by_date(ws, dateFrom, dateTo)
    save(dstFileName, wb)    

def main():
    parser = OptionParser()
    parser.add_option("-i",
                      "--input",
                      dest="srcFileName",
                      type=str, 
                      help="ideabox xlsx file")
    parser.add_option("-o",
                      "--output",
                      dest="dstFileName",
                      type=str,
                      help="parsed xlsx file")
    parser.add_option(
                      "--from",
                      dest="fromDate",
                      help="%Y-%m-%d format(e.g. 2018-10-10), idea submitted before the date will be discarded.")
    parser.add_option(
                      "--end",
                      dest="endDate",
                      help="%Y-%m-%d format(e.g. 2018-10-10), idea submitted after the date will be discarded")
    parser.add_option("--debug",
                      dest="verbose",
                      action="store_true",
                      default=False,
                      help="print more info")

    options, args = parser.parse_args()
    
    # not working
    #if options.verbose == True:
    #    debugLevel = 'INFO'


    if not options.srcFileName:
        parser.error("Missing mandatory argument: '-i'")
    if not options.fromDate:
        parser.error("Missing mandatory argument: '--from'")    

    try:
        dateFrom = datetime.datetime.strptime(options.fromDate,"%Y-%m-%d")            
        if not options.endDate:            
            dateTo = datetime.datetime.now()            
        else:
            dateTo = datetime.datetime.strptime(options.endDate,"%Y-%m-%d")
    except Exception:        
        parser.error("date is not correct")

    srcFileName = options.srcFileName
    if not options.dstFileName:
        [realFileName, fileSuffix] = srcFileName.split('.')
        dstFileName = realFileName + '_' + dateFrom.strftime("%Y%m%d") + '_' + dateTo.strftime("%Y%m%d") + '.' + fileSuffix

    parse_idea_box(srcFileName,dstFileName,dateFrom,dateTo)            

if __name__ == "__main__":
    main()
